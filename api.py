from flask import Flask, request, jsonify, send_file, url_for
import mysql.connector
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

app = Flask(__name__)

@app.route("/api/query", methods=["POST"])
def executar_query():
    data = request.get_json()
    query = data.get("sql")

    if not query:
        return "A consulta não foi fornecida", 400

    conn = mysql.connector.connect(
        host="xxxxxxxxxxxr",
        user="xxxxxxxxxxxx",
        password="xxxxxxxxx',
        database="xxxxxxxxxx"
    )
    cursor = conn.cursor()
    cursor.execute(query)
    resultados = cursor.fetchall()
    cursor.close()
    conn.close()

    nome_arquivo_excel = criar_excel(resultados)
    download_url = url_for("download_file", filename=nome_arquivo_excel, _external=True)

    return jsonify({"filename": nome_arquivo_excel, "download_url": download_url})


@app.route("/api/download/<filename>", methods=["GET"])
def download_file(filename):
    return send_file(filename, as_attachment=True)


def criar_excel(resultados):
    contador_sequencial = obter_contador_sequencial()
    nome_arquivo_excel = f"DEBITO ({contador_sequencial}).xlsx"
    contador_sequencial += 1
    salvar_contador_sequencial(contador_sequencial)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Débito"

    headers = ["NOTA", "FORNECEDOR", "EMISSAO", "VENCIMENTO", "VALOR", "HISTÓRICO"]
    sheet.append(headers)

    header_font = Font(bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    border = Border(bottom=Side(border_style="thin"))
    fill = PatternFill(fill_type="solid", fgColor="4A8FE7")  # Cor azul

    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = alignment
        cell.border = border
        cell.fill = fill

    valores = []  # Lista para armazenar os valores para cálculo do débito total

    for resultado in resultados:
        resultado_sem_simbolo = [valor[2:].replace(".", "").replace(",", ".") if isinstance(valor, str) and valor.startswith("R$") else valor for valor in resultado]
        sheet.append(resultado_sem_simbolo)
        valores.append(float(resultado_sem_simbolo[4]))  # Adiciona o valor à lista de valores

    valor_column = sheet["E"]
    valor_alignment = Alignment(horizontal="right", vertical="center")
    valor_border = Border(right=Side(border_style="thin"))

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=valor_column[0].column, max_col=valor_column[-1].column):
        for cell in row:
            cell.alignment = valor_alignment
            cell.border = valor_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"

    total_row = sheet.max_row + 1
    total_cell = sheet.cell(row=total_row, column=1, value="Débito Total:")
    total_cell.font = header_font
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.fill = fill

    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        column_letter = column[0].column_letter
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width

    total_cell.border = Border(top=Side(border_style="thin"))

    valor_column = [cell for cell in sheet["E"] if isinstance(cell.value, (int, float))]
    total_value = sum(valores)  # Soma os valores da lista
    total_value_cell = sheet.cell(row=total_row, column=valor_column[0].column if valor_column else 5, value=total_value)
    total_value_cell.number_format = "#,##0.00"
    total_value_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_value_cell.fill = fill

    try:
        workbook.save(nome_arquivo_excel)
    except Exception as e:
        print(f"Erro ao salvar o arquivo: {e}")

    return nome_arquivo_excel


def obter_contador_sequencial():
    arquivo_contador = "contador_sequencial.txt"
    if os.path.exists(arquivo_contador):
        with open(arquivo_contador, "r") as file:
            contador = int(file.read())
    else:
        contador = 1
    return contador


def salvar_contador_sequencial(contador):
    arquivo_contador = "contador_sequencial.txt"
    with open(arquivo_contador, "w") as file:
        file.write(str(contador))


if __name__ == "__main__":
    app.run(host='0.0.0.0')
